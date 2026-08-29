"""
reconcile.py — Bank-to-GL three-pass reconciliation engine.

Author: Raghav Khanna — BBA (Co-op), Goodman School of Business, Brock University
        CPA-track · linkedin.com/in/raghav-khanna-73535932a

Reads a bank statement CSV and a GL cash-account extract CSV, matches them in
three passes of decreasing strictness, classifies whatever is left as
exceptions with a probable cause and a suggested action, and writes a
multi-tab, close-ready Excel workbook.

    Pass 1 — EXACT      same signed amount, same date
    Pass 2 — TIMING     same signed amount, dates within ±5 days
                        (checks clearing late, EFT settlement lag)
    Pass 3 — TOLERANCE  amounts within $0.99, dates within ±7 days,
                        fuzzy description similarity (rounding / keying errors)

Everything unmatched after pass 3 is an exception. Exceptions are ranked by
dollar exposure and classified:

    bank side:  bank charge / interest / NSF not booked -> book a JE
                unidentified debit or credit            -> investigate
    GL side:    outstanding check                       -> carry on rec
                deposit in transit                      -> verify July clear
                duplicate posting                       -> reverse the entry

Usage:
    python src/reconcile.py \
        --bank data/bank_statement_jun2026.csv \
        --gl   data/gl_cash_extract_jun2026.csv \
        --out  output/reconciliation_report_jun2026.xlsx
"""

import argparse
import re
from datetime import date
from difflib import SequenceMatcher
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ------------------------------------------------------------------ parameters
TIMING_WINDOW_DAYS = 5      # pass 2: max days between GL booking and bank clearing
TOLERANCE_DOLLARS = 0.99    # pass 3: max absolute amount difference
TOLERANCE_WINDOW_DAYS = 7   # pass 3: max days apart
FUZZY_THRESHOLD = 0.35      # pass 3: min description similarity (0..1)
OPENING_BALANCE = 184_352.19
PERIOD_LABEL = "June 2026"
ENTITY_LABEL = "Operating Account 1010 — Cash"

STOPWORDS = {"EFT", "PAD", "POS", "DEP", "CHQ", "ADP", "PPD", "INC", "LTD",
             "SVC", "PMT", "PAYMENT", "INVOICE", "THE", "OF", "AND", "CDA",
             "CANADA", "REF"}


# ------------------------------------------------------------------ helpers
def normalize(desc: str) -> str:
    s = re.sub(r"[^A-Z0-9 ]", " ", str(desc).upper())
    s = re.sub(r"\b(REF|CHQ)?\d+\b", " ", s)          # strip reference numbers
    tokens = [t for t in s.split() if t not in STOPWORDS and len(t) > 2]
    return " ".join(tokens)


def similarity(a: str, b: str) -> float:
    """Blend of sequence ratio and token overlap — robust to reordering."""
    na, nb = normalize(a), normalize(b)
    if not na or not nb:
        return 0.0
    seq = SequenceMatcher(None, na, nb).ratio()
    ta, tb = set(na.split()), set(nb.split())
    jac = len(ta & tb) / len(ta | tb) if ta | tb else 0.0
    return max(seq, jac)


def load(bank_path: str, gl_path: str):
    bank = pd.read_csv(bank_path, parse_dates=["Date"])
    gl = pd.read_csv(gl_path, parse_dates=["Date"])
    bank["Amount"] = bank["Amount"].astype(float)
    gl["Amount"] = gl["Amount"].astype(float)
    bank["_key"] = (bank["Amount"] * 100).round().astype(int)  # cents key, no float fuzz
    gl["_key"] = (gl["Amount"] * 100).round().astype(int)
    return bank.reset_index(drop=True), gl.reset_index(drop=True)


# ------------------------------------------------------------------ matching
def run_matching(bank: pd.DataFrame, gl: pd.DataFrame):
    """Greedy one-to-one matching in three passes. Returns (matches, bank_open, gl_open)."""
    matches = []          # dicts with bank idx, gl idx, pass name, deltas, score
    bank_open = set(bank.index)
    gl_open = set(gl.index)

    def pair(bi, gi, pass_name, score=None):
        b, g = bank.loc[bi], gl.loc[gi]
        matches.append({
            "pass": pass_name,
            "bank_idx": bi, "gl_idx": gi,
            "date_delta": int((b["Date"] - g["Date"]).days),
            "amount_delta": round(b["Amount"] - g["Amount"], 2),
            "score": score,
        })
        bank_open.discard(bi)
        gl_open.discard(gi)

    # Pass 1 — exact: same cents, same date
    gl_by_key = {}
    for gi in gl_open:
        gl_by_key.setdefault(gl.at[gi, "_key"], []).append(gi)
    for bi in sorted(bank_open):
        key = bank.at[bi, "_key"]
        for gi in gl_by_key.get(key, []):
            if gi in gl_open and gl.at[gi, "Date"] == bank.at[bi, "Date"]:
                pair(bi, gi, "Exact")
                break

    # Pass 2 — timing: same cents, nearest date within the window
    for bi in sorted(bank_open):
        key = bank.at[bi, "_key"]
        best, best_dd = None, None
        for gi in gl_by_key.get(key, []):
            if gi not in gl_open:
                continue
            dd = abs((bank.at[bi, "Date"] - gl.at[gi, "Date"]).days)
            if dd <= TIMING_WINDOW_DAYS and (best is None or dd < best_dd):
                best, best_dd = gi, dd
        if best is not None:
            pair(bi, best, "Timing")

    # Pass 3 — tolerance + fuzzy description: best similarity wins
    tol_cents = int(TOLERANCE_DOLLARS * 100)
    for bi in sorted(bank_open):
        b = bank.loc[bi]
        best, best_score = None, 0.0
        for gi in sorted(gl_open):
            g = gl.loc[gi]
            if abs(b["_key"] - g["_key"]) > tol_cents or b["_key"] == g["_key"]:
                continue  # same-cents pairs belong to passes 1-2
            if abs((b["Date"] - g["Date"]).days) > TOLERANCE_WINDOW_DAYS:
                continue
            score = similarity(b["Description"], g["Memo"])
            if score >= FUZZY_THRESHOLD and score > best_score:
                best, best_score = gi, score
        if best is not None:
            pair(bi, best, "Tolerance", round(best_score, 2))

    return matches, bank_open, gl_open


# ------------------------------------------------------------------ exceptions
FEE_PATTERN = re.compile(r"\bFEE\b|SERVICE CHARGE|SVC CHG|\bNSF\b|OVERDRAFT", re.I)
INTEREST_PATTERN = re.compile(r"\bINTEREST\b", re.I)

MONTH_END_DAYS = 5  # "near month end" = within the last N days of the period


def classify_exceptions(bank, gl, bank_open, gl_open, matches):
    period_end = max(bank["Date"].max(), gl["Date"].max())
    matched_gl = {(gl.at[m["gl_idx"], "_key"], normalize(gl.at[m["gl_idx"], "Memo"]))
                  for m in matches}
    rows = []

    for bi in sorted(bank_open):
        b = bank.loc[bi]
        if INTEREST_PATTERN.search(b["Description"]):
            cause = "Interest earned — not booked in GL"
            action = "Book JE: DR 1010 Cash / CR 4210 Interest Income"
        elif FEE_PATTERN.search(b["Description"]):
            cause = "Bank charge — not booked in GL"
            action = "Book JE: DR 6220 Bank Charges / CR 1010 Cash"
        elif b["Amount"] > 0:
            cause = "Unidentified bank credit"
            action = "Trace with bank / AR — identify payer before booking"
        else:
            cause = "Unidentified bank debit"
            action = "Investigate with bank — possible unauthorized PAD"
        rows.append({
            "Side": "Bank only", "Date": b["Date"], "Description": b["Description"],
            "Doc/Ref": b["Reference"], "Amount": b["Amount"],
            "Probable cause": cause, "Suggested action": action,
        })

    for gi in sorted(gl_open):
        g = gl.loc[gi]
        near_eom = (period_end - g["Date"]).days <= MONTH_END_DAYS
        is_dup = (g["_key"], normalize(g["Memo"])) in matched_gl
        if is_dup:
            cause = "Possible duplicate posting (identical to a matched entry)"
            action = "Review source document — reverse the duplicate JE"
        elif g["Amount"] < 0 and str(g["DocNo"]).startswith("CHQ"):
            cause = "Outstanding check" + (" (issued near month end)" if near_eom else "")
            action = "Carry as reconciling item — follow up if stale > 60 days"
        elif g["Amount"] < 0:
            cause = "Payment initiated, not yet cleared"
            action = "Confirm settlement in July statement"
        else:
            cause = "Deposit in transit"
            action = "Verify credit on July bank statement"
        rows.append({
            "Side": "GL only", "Date": g["Date"], "Description": g["Memo"],
            "Doc/Ref": g["DocNo"], "Amount": g["Amount"],
            "Probable cause": cause, "Suggested action": action,
        })

    rows.sort(key=lambda r: abs(r["Amount"]), reverse=True)
    for rank, r in enumerate(rows, 1):
        r["Rank"] = rank
    return rows


# ------------------------------------------------------------------ excel report
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F3864")
SUB_FONT = Font(size=10, color="595959")
MONEY = '#,##0.00;[Red](#,##0.00)'

CAUSE_FILLS = {
    "duplicate": PatternFill("solid", fgColor="F8CBAD"),
    "Unidentified": PatternFill("solid", fgColor="FFC7CE"),
    "Outstanding": PatternFill("solid", fgColor="FFF2CC"),
    "Deposit in transit": PatternFill("solid", fgColor="DDEBF7"),
    "Bank charge": PatternFill("solid", fgColor="E2EFDA"),
    "Interest": PatternFill("solid", fgColor="E2EFDA"),
}


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(vertical="center")
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_table(ws, start_row, headers, rows, money_cols, date_cols):
    ws.append([])  # ensure openpyxl row cursor sane when using explicit cells
    for c, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=c, value=h)
    style_header(ws, start_row, len(headers))
    r = start_row
    for rowdata in rows:
        r += 1
        for c, v in enumerate(rowdata, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = BORDER
            if c in money_cols:
                cell.number_format = MONEY
            if c in date_cols:
                cell.number_format = "yyyy-mm-dd"
    return r


def build_report(bank, gl, matches, exceptions, out_path):
    wb = Workbook()

    m_by_pass = {"Exact": [], "Timing": [], "Tolerance": []}
    for m in matches:
        m_by_pass[m["pass"]].append(m)

    bank_total = bank["Amount"].sum()
    gl_total = gl["Amount"].sum()
    bank_end = OPENING_BALANCE + bank_total
    gl_end = OPENING_BALANCE + gl_total

    exc = {"dit": 0.0, "oschk": 0.0, "unbooked": 0.0, "dup": 0.0, "unid": 0.0}
    for e in exceptions:
        cause = e["Probable cause"]
        if "Deposit in transit" in cause:
            exc["dit"] += e["Amount"]
        elif "Outstanding check" in cause or "not yet cleared" in cause:
            exc["oschk"] += e["Amount"]
        elif "duplicate" in cause:
            exc["dup"] += e["Amount"]
        elif "Unidentified" in cause:
            exc["unid"] += e["Amount"]
        else:
            exc["unbooked"] += e["Amount"]

    residual = round(sum(m["amount_delta"] for m in matches if m["pass"] == "Tolerance"), 2)
    adj_bank = bank_end + exc["dit"] + exc["oschk"]          # bank +DIT -O/S checks
    adj_gl = gl_end + exc["unbooked"] - exc["dup"] + exc["unid"] + residual

    # ---------------- Summary tab
    ws = wb.active
    ws.title = "Summary"
    ws.sheet_properties.tabColor = "1F3864"
    ws["A1"] = "Bank-to-GL Reconciliation"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = f"{ENTITY_LABEL}  ·  Period: {PERIOD_LABEL}"
    ws["A2"].font = SUB_FONT
    ws["A3"] = "Prepared by: Raghav Khanna  ·  Engine: three-pass matcher (exact / timing / tolerance+fuzzy)"
    ws["A3"].font = SUB_FONT

    stats = [
        ("Bank transactions", len(bank)),
        ("GL transactions", len(gl)),
        ("Matched — Pass 1 (exact)", len(m_by_pass["Exact"])),
        ("Matched — Pass 2 (timing ±5 days)", len(m_by_pass["Timing"])),
        ("Matched — Pass 3 (tolerance ≤ $0.99 + fuzzy)", len(m_by_pass["Tolerance"])),
        ("Total reconciled", len(matches)),
        ("Match rate (bank side)", f"{len(matches) / len(bank):.1%}"),
        ("Exceptions isolated", len(exceptions)),
        ("Exception exposure (gross $)", round(sum(abs(e["Amount"]) for e in exceptions), 2)),
    ]
    r = write_table(ws, 5, ["Metric", "Value"], stats, money_cols={2}, date_cols=set())
    ws.cell(row=11, column=2).number_format = "0"      # total reconciled as int
    for rr in (6, 7, 8, 9, 10, 12):
        ws.cell(row=rr, column=2).number_format = "0"

    proof = [
        ("Ending balance per bank statement", round(bank_end, 2)),
        ("  add: deposits in transit", round(exc["dit"], 2)),
        ("  less: outstanding checks / uncleared payments", round(exc["oschk"], 2)),
        ("Adjusted bank balance", round(adj_bank, 2)),
        ("Ending balance per general ledger", round(gl_end, 2)),
        ("  add: bank charges / interest not booked (net)", round(exc["unbooked"], 2)),
        ("  add back: duplicate posting to reverse", round(-exc["dup"], 2)),
        ("  add: unidentified bank items (net, pending ID)", round(exc["unid"], 2)),
        ("  add: pass-3 amount residuals (pending write-off JE)", residual),
        ("Adjusted GL balance", round(adj_gl, 2)),
        ("Unreconciled difference", round(adj_bank - adj_gl, 2)),
    ]
    last_proof = write_table(ws, r + 2, ["Reconciliation proof", "Amount"], proof,
                             money_cols={2}, date_cols=set())
    for rr in range(r + 3, last_proof + 1):
        label = str(ws.cell(row=rr, column=1).value)
        if label.startswith(("Adjusted", "Unreconciled", "Ending")):
            ws.cell(row=rr, column=1).font = Font(bold=True)
            ws.cell(row=rr, column=2).font = Font(bold=True)
    autosize(ws, [52, 20])

    # ---------------- Exceptions tab
    ws = wb.create_sheet("Exceptions")
    ws.sheet_properties.tabColor = "C00000"
    ws["A1"] = "Exception Report — ranked by dollar exposure"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = "Each break carries a probable cause and the action that clears it."
    ws["A2"].font = SUB_FONT
    headers = ["Rank", "Side", "Date", "Description", "Doc/Ref", "Amount",
               "Probable cause", "Suggested action"]
    rows = [[e["Rank"], e["Side"], e["Date"], e["Description"], e["Doc/Ref"],
             e["Amount"], e["Probable cause"], e["Suggested action"]]
            for e in exceptions]
    last = write_table(ws, 4, headers, rows, money_cols={6}, date_cols={3})
    for rr in range(5, last + 1):
        cause = str(ws.cell(row=rr, column=7).value)
        for key, fill in CAUSE_FILLS.items():
            if key.lower() in cause.lower():
                for cc in range(1, 9):
                    ws.cell(row=rr, column=cc).fill = fill
                break
    autosize(ws, [6, 10, 12, 44, 12, 14, 46, 46])

    # ---------------- Matched tabs
    def matched_tab(name, color, mlist, note):
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = color
        ws["A1"] = note
        ws["A1"].font = SUB_FONT
        headers = ["Bank date", "Bank description", "Bank amount",
                   "GL date", "GL memo", "GL doc", "GL amount",
                   "Days Δ", "Amount Δ", "Match score"]
        rows = []
        for m in sorted(mlist, key=lambda x: bank.at[x["bank_idx"], "Date"]):
            b, g = bank.loc[m["bank_idx"]], gl.loc[m["gl_idx"]]
            rows.append([b["Date"], b["Description"], b["Amount"],
                         g["Date"], g["Memo"], g["DocNo"], g["Amount"],
                         m["date_delta"], m["amount_delta"],
                         m["score"] if m["score"] is not None else "—"])
        write_table(ws, 3, headers, rows, money_cols={3, 7, 9}, date_cols={1, 4})
        autosize(ws, [12, 38, 14, 12, 38, 12, 14, 8, 11, 12])

    matched_tab("Matched — Exact", "375623", m_by_pass["Exact"],
                "Pass 1: identical signed amount and identical date.")
    matched_tab("Matched — Timing", "548235", m_by_pass["Timing"],
                f"Pass 2: identical amount, cleared within ±{TIMING_WINDOW_DAYS} days (checks/EFT settlement lag).")
    matched_tab("Matched — Tolerance", "70AD47", m_by_pass["Tolerance"],
                f"Pass 3: amount within ${TOLERANCE_DOLLARS}, ±{TOLERANCE_WINDOW_DAYS} days, "
                f"description similarity ≥ {FUZZY_THRESHOLD}. Amount Δ = residual to write off or adjust.")

    # ---------------- Source tabs
    def source_tab(name, df, cols, money_col):
        ws = wb.create_sheet(name)
        ws.sheet_properties.tabColor = "808080"
        rows = df[cols].values.tolist()
        write_table(ws, 1, cols, rows, money_cols={money_col}, date_cols={1})
        autosize(ws, [12] + [34] * (len(cols) - 2) + [14])

    source_tab("Bank Statement (source)", bank.sort_values("Date"),
               ["Date", "Description", "Reference", "Amount"], 4)
    source_tab("GL Extract (source)", gl.sort_values("Date"),
               ["Date", "Account", "Memo", "DocNo", "Amount"], 5)

    out_path = Path(out_path)
    out_path.parent.mkdir(exist_ok=True)
    wb.save(out_path)
    return out_path


# ------------------------------------------------------------------ main
def main():
    ap = argparse.ArgumentParser(description="Three-pass bank-to-GL reconciliation")
    root = Path(__file__).resolve().parent.parent
    ap.add_argument("--bank", default=str(root / "data/bank_statement_jun2026.csv"))
    ap.add_argument("--gl", default=str(root / "data/gl_cash_extract_jun2026.csv"))
    ap.add_argument("--out", default=str(root / "output/reconciliation_report_jun2026.xlsx"))
    args = ap.parse_args()

    bank, gl = load(args.bank, args.gl)
    matches, bank_open, gl_open = run_matching(bank, gl)
    exceptions = classify_exceptions(bank, gl, bank_open, gl_open, matches)
    out = build_report(bank, gl, matches, exceptions, args.out)

    n = {"Exact": 0, "Timing": 0, "Tolerance": 0}
    for m in matches:
        n[m["pass"]] += 1
    print(f"Bank rows: {len(bank)}   GL rows: {len(gl)}")
    print(f"Pass 1 exact:     {n['Exact']:>3}")
    print(f"Pass 2 timing:    {n['Timing']:>3}")
    print(f"Pass 3 tolerance: {n['Tolerance']:>3}")
    print(f"Total reconciled: {len(matches):>3}")
    print(f"Exceptions:       {len(exceptions):>3}  "
          f"(bank-only {len(bank_open)}, GL-only {len(gl_open)})")
    print(f"Report: {out}")


if __name__ == "__main__":
    main()
